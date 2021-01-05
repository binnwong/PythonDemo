# coding=utf-8
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment
import time


openpyxl_data = [
    ('我', '们', '在', '这', '寻', '找'),
    ('我', '们', '在', '这', '失', '去'),
    ('p', 'y', 't', 'h', 'o', 'n')
]
output_file_name = 'openpyxl_file.xlsx'


def save_excel(target_list, output_file_name):
    """
    将数据写入xlsx文件
    """
    if not output_file_name.endswith('.xlsx'):
        output_file_name += '.xlsx'

    # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
    wb = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    title_data = ('a', 'b', 'c', 'd', 'e', 'f')
    target_list.insert(0, title_data)
    rows = len(target_list)
    lines = len(target_list[0])
    for i in range(rows):
        for j in range(lines):
            ws.cell(row=i + 1, column=j + 1).value = target_list[i][j]

    # 获取每一列的内容的最大宽度
    i = 0
    col_width = list()
    # 每列
    for col in ws.columns:
        # 每行
        for j in range(len(col)):
            if j == 0:
                # 数组增加一个元素
                col_width.append(len(str(col[j].value)))
            else:
                # 获得每列中的内容的最大宽度
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1
    # print(col_width)
    # 设置列宽
    for i in range(len(col_width)):
        # 根据列的数字返回字母
        col_letter = get_column_letter(i + 1)
        # 当宽度大于40，宽度设置为45
        if col_width[i] > 40:
            ws.column_dimensions[col_letter].width = 45
            # for j in range(len(list(ws.columns)[i])):
            #     if len(list(ws.columns)[i][j].value) > 40:
            #         ws.row_dimensions[j].height = 15 * (len(list(ws.columns)[i][j].value) // 40 + 1)
        else:
            ws.column_dimensions[col_letter].width = col_width[i] + 3
    # 设置单元格对齐格式
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=10)

    # 设置到期日期超过当前时间的值为红色字体
    for col in ws.columns:
        if col[0].value == '到期日期':
            for j in range(1, len(col)):
                if col[j].value and col[j].value != 'None':
                    due_time = time.mktime(time.strptime(col[j].value, '%Y-%m-%d'))
                    local_time = time.mktime(time.localtime())
                    if int(due_time) <= int(local_time):
                        # print('due_time', int(due_time), 'local_time', int(local_time))
                        col[j].font = Font(color=colors.RED)
    # 保存表格
    wb.save(filename=output_file_name)


save_excel(openpyxl_data, output_file_name)


# import openpyxl
#
#
# input_file_name = 'openpyxl_file.xlsx'
#
#
# def read_excel(input_file_name):
#     """
#     从xlsx文件中读取数据
#     """
#     workbook = openpyxl.load_workbook(input_file_name)
#     print(workbook)
#     # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
#     print(workbook.sheetnames)
#     table = workbook.active
#     print(table)
#     rows = table.max_row
#     cols = table.max_column
#
#     for row in range(rows):
#         for col in range(cols):
#             data = table.cell(row + 1, col + 1).value
#             print(data, end=' ')
#
#
# read_excel(input_file_name)
